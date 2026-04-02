import pandas as pd
import pytesseract
import re
import os
import sys
from pdf2image import convert_from_path
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

if sys.platform == 'win32':
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


def read_master(master_path):
    """Citeste MASTER si returneaza lista de dictionare cu datele per pozitie."""
    xls = pd.ExcelFile(master_path)
    all_entries = []
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=0)
        cols = {str(c).strip().upper(): i for i, c in enumerate(df.columns)}

        for i in range(len(df)):
            row = df.iloc[i]
            nr_pv = row.iloc[0]
            if pd.isna(nr_pv):
                continue
            try:
                nr_pv_int = int(float(nr_pv))
            except (ValueError, TypeError):
                continue

            def get_val(col_idx, default=''):
                v = row.iloc[col_idx] if col_idx < len(row) else default
                return str(v).strip() if pd.notna(v) else default

            entry = {
                'sheet': sheet_name,
                'nr_pv': nr_pv_int,
                'pozitie_hg': get_val(1),
                'judet': get_val(2),
                'uat': get_val(3),
                'nume1': get_val(5),
                'adresa1': get_val(6),
                'tarla': get_val(11),
                'parcela': get_val(12),
                'nr_cadastral': get_val(13),
                'nr_cf': get_val(14),
                'categorie': get_val(15),
                'extravilan_intravilan': get_val(16),
                'suprafata_totala': get_val(17),
                'suprafata_exp1': get_val(18),
                'valoare1': get_val(19),
                'suprafata_exp2': get_val(20),
                'valoare2': get_val(21),
                'hg': get_val(28),
                'decizie_expropriere': get_val(31),
                'decizie_comisie': get_val(32),
                'membru1': get_val(33),
                'membru2': get_val(34),
                'membru3': get_val(35),
                'membru4': get_val(36),
                'membru5': get_val(37),
                'rezulta': get_val(39),
                'data': get_val(40),
                'ora': get_val(41),
            }
            all_entries.append(entry)
    return all_entries


def ocr_pdf(pdf_path):
    """Extrage text din PDF (scanat sau text)."""
    # Try text extraction first
    reader = PdfReader(pdf_path)
    full_text = ''
    for page in reader.pages:
        t = page.extract_text()
        if t:
            full_text += t + '\n'

    if len(full_text.strip()) > 200:
        return full_text

    # OCR fallback
    images = convert_from_path(pdf_path, dpi=250)
    texts = []
    for img in images:
        t = pytesseract.image_to_string(img, lang='ron')
        texts.append(t)
    return '\n'.join(texts)


def extract_data_from_pdf(text, filename):
    """Extrage date structurate din textul OCR al unui H si PV."""
    data = {'filename': filename, 'raw_text_length': len(text)}

    # Nr. HSD/PV si data
    m = re.search(r'nr\.\s*(\d+)\s*din\s*([\d.]+\.\d{4})', text, re.IGNORECASE)
    if m:
        data['nr_pv'] = int(m.group(1))
        data['data'] = m.group(2)

    # UAT
    m = re.search(r'CONSILIUL\s+LOCAL\s+([A-ZĂÂÎȘȚ\-\s]+?)\s*,\s*JUDE[ŢȚ]UL\s+([A-ZĂÂÎȘȚ]+)', text)
    if m:
        data['uat'] = m.group(1).strip()
        data['judet'] = m.group(2).strip()

    # Pozitie HG
    m = re.search(r'pozi[tț]i[ae]\s*(?:nr\.?)?\s*(\d+)\s*din\s*Anex', text, re.IGNORECASE)
    if m:
        data['pozitie_hg'] = m.group(1)

    # Nr cadastral
    m = re.search(r'num[aă]r\s+cadastral\s*/?\s*(?:nr\.?)?\s*topo\s+(\d+)', text, re.IGNORECASE)
    if m:
        data['nr_cadastral'] = m.group(1)

    # Nr CF
    m = re.search(r'carte\s+funciar[aă]\s+(\d+)', text, re.IGNORECASE)
    if m:
        data['nr_cf'] = m.group(1)

    # Tarla si parcela
    m = re.search(r'tarla\s*(?:nr\.?)?\s*([\d/]+)\s*,?\s*parcela\s*(?:nr\.?)?\s*([\d/]+)', text, re.IGNORECASE)
    if m:
        data['tarla'] = m.group(1)
        data['parcela'] = m.group(2)

    # Suprafata expropriata
    surfaces = re.findall(r'Teren\d?\s+[iî]n\s+suprafa[tț][aă]\s+de\s+([\d.,]+)\s*mp', text, re.IGNORECASE)
    if surfaces:
        data['suprafete_pdf'] = [s.replace('.', '').replace(',', '.') for s in surfaces]

    # Valoare despagubiri
    vals = re.findall(r'([\d.,]+)\s*LEI\s+pentru\s+imobilul\s+teren', text, re.IGNORECASE)
    if vals:
        data['valori_pdf'] = [v.replace('.', '').replace(',', '.') for v in vals]

    # Suma totala
    m = re.search(r'suma\s+de\s+([\d.,]+)\s*LEI', text, re.IGNORECASE)
    if m:
        data['suma_totala_pdf'] = m.group(1).replace('.', '').replace(',', '.')

    # Proprietar - multiple strategies
    prop_name = None
    # Strategy 1: after "supus exproprierii" + numbered list
    m1 = re.search(r'supus[eă]?\s+exproprierii.*?\n\s*1[\.\)]\s*(.+?)(?:,\s*cu\s+(?:domiciliul|sediul)|;\s*\n)', text, re.IGNORECASE | re.DOTALL)
    if m1:
        prop_name = re.sub(r'\s+', ' ', m1.group(1)).strip()
    # Strategy 2: "Art. 1" section with proprietar
    if not prop_name:
        m2 = re.search(r'Art\.\s*1\..*?localitat.*?(?:1[\.\)]\s*)?([A-ZĂÂÎȘȚÜÖ][A-ZĂÂÎȘȚÜÖ\s\-\.]+(?:SRL|SA|S\.R\.L\.|S\.A\.)?)\s*,?\s*(?:cu\s+(?:domiciliul|sediul)|$)', text, re.DOTALL)
        if m2:
            prop_name = m2.group(1).strip()
    # Filter out false positives
    if prop_name and any(kw in prop_name for kw in ['REPREZENTANT', 'Primaria', 'AV.', 'LEI', 'teren']):
        prop_name = None
    if prop_name:
        data['proprietar_pdf'] = prop_name

    # Decizie expropriere
    m = re.search(r'[Dd]eciziei\s+de\s+expropriere\s+nr\.\s*(\d+)\s+din\s+([\d.]+)', text)
    if m:
        data['decizie_expropriere_pdf'] = f'nr. {m.group(1)} din {m.group(2)}'

    # Decizie comisie
    m = re.search(r'[Dd]eciziei\s+nr\.\s*(\d+)\s*\n?\s*din\s+([\d.]+)\s*(?:\d+\s*)?emis', text)
    if m:
        data['decizie_comisie_pdf'] = f'nr. {m.group(1)} din {m.group(2)}'

    # HG number - look for the specific pattern "H.G. nr. XXXX/YYYY" or "Guvernului nr. XXXX/YYYY"
    m = re.search(r'(?:Guvernului|H\.?\s*G\.?)\s*nr\.?\s*(\d{2,4}\s*/\s*\d{4})', text)
    if m:
        data['hg_pdf'] = re.sub(r'\s', '', m.group(1))

    # Membri comisie - cautam numele dupa REPREZENTANT
    membri = re.findall(r'REPREZENTANT.*?\s{2,}([A-ZĂÂÎȘȚ][a-zăâîșț]+(?:\s+[A-ZĂÂÎȘȚ\-][a-zăâîșțA-ZĂÂÎȘȚ\-]+)+)', text)
    if membri:
        data['membri_pdf'] = [m.strip() for m in membri]
    # Si AV.
    avocati = re.findall(r'AV\.\s+([A-ZĂÂÎȘȚ][A-ZĂÂÎȘȚ\-\s]+)', text)
    if avocati:
        if 'membri_pdf' not in data:
            data['membri_pdf'] = []
        data['membri_pdf'].extend([a.strip() for a in avocati])

    # Rezulta / Nu rezulta
    if 'nu a fost depus' in text.lower() or 'nu rezulta' in text.lower():
        data['rezulta_pdf'] = 'NU REZULTA'
    elif 'rezulta' in text.lower():
        data['rezulta_pdf'] = 'REZULTA'

    return data


def compare_entry(master, pdf):
    """Compara datele din MASTER cu cele din PDF. Returneaza lista de neconcordante."""
    issues = []

    def normalize(s):
        return re.sub(r'\s+', ' ', str(s).strip().upper().replace('Ț', 'T').replace('Ş', 'S').replace('Ă', 'A').replace('Â', 'A').replace('Î', 'I'))

    def check(field_name, master_val, pdf_val):
        if not pdf_val:
            return  # nu am reusit sa extrag din PDF, nu raportam
        m = normalize(master_val)
        p = normalize(pdf_val)
        if m and p and m != p:
            # Check partial match for names and long strings
            if m in p or p in m:
                return
            issues.append({
                'camp': field_name,
                'master': str(master_val).strip(),
                'pdf': str(pdf_val).strip(),
            })

    # UAT
    check('UAT', master.get('uat', ''), pdf.get('uat', ''))

    # Judet
    check('Județ', master.get('judet', ''), pdf.get('judet', ''))

    # Pozitie HG
    check('Poziție HG', master.get('pozitie_hg', ''), pdf.get('pozitie_hg', ''))

    # Nr cadastral
    check('Nr. cadastral', master.get('nr_cadastral', ''), pdf.get('nr_cadastral', ''))

    # Nr CF
    check('Nr. CF', master.get('nr_cf', ''), pdf.get('nr_cf', ''))

    # Tarla
    check('Tarla', master.get('tarla', ''), pdf.get('tarla', ''))

    # Parcela
    check('Parcela', master.get('parcela', ''), pdf.get('parcela', ''))

    # Data
    check('Data', master.get('data', ''), pdf.get('data', ''))

    # Proprietar
    if pdf.get('proprietar_pdf'):
        m_name = normalize(master.get('nume1', ''))
        p_name = normalize(pdf.get('proprietar_pdf', ''))
        if m_name and p_name:
            # Check if any significant part matches
            m_parts = set(m_name.split())
            p_parts = set(p_name.split())
            common = m_parts & p_parts
            if len(common) < min(2, len(m_parts)):
                issues.append({
                    'camp': 'Proprietar',
                    'master': master.get('nume1', ''),
                    'pdf': pdf.get('proprietar_pdf', ''),
                })

    # Suprafete
    if pdf.get('suprafete_pdf'):
        master_sups = []
        for key in ['suprafata_exp1', 'suprafata_exp2']:
            v = master.get(key, '')
            if v and v != '' and v != '0':
                try:
                    master_sups.append(float(str(v).replace(',', '.')))
                except ValueError:
                    pass
        pdf_sups = []
        for s in pdf['suprafete_pdf']:
            try:
                pdf_sups.append(float(s))
            except ValueError:
                pass
        if master_sups and pdf_sups:
            for ms in master_sups:
                if not any(abs(ms - ps) < 1 for ps in pdf_sups):
                    issues.append({
                        'camp': 'Suprafață expropriată',
                        'master': str(int(ms)),
                        'pdf': ', '.join(str(int(p)) for p in pdf_sups),
                    })
                    break

    # Valori
    if pdf.get('valori_pdf') or pdf.get('suma_totala_pdf'):
        master_vals = []
        for key in ['valoare1', 'valoare2']:
            v = master.get(key, '')
            if v and v != '' and v != '0':
                try:
                    master_vals.append(float(str(v).replace(',', '.')))
                except ValueError:
                    pass
        master_total = sum(master_vals)
        pdf_total = None
        if pdf.get('suma_totala_pdf'):
            try:
                pdf_total = float(pdf['suma_totala_pdf'])
            except ValueError:
                pass
        if master_total > 0 and pdf_total and abs(master_total - pdf_total) > 1:
            issues.append({
                'camp': 'Valoare despăgubiri totală',
                'master': f'{master_total:,.2f}'.replace(',', '.'),
                'pdf': f'{pdf_total:,.2f}'.replace(',', '.'),
            })

    # Decizie expropriere
    if pdf.get('decizie_expropriere_pdf'):
        m_dec = normalize(master.get('decizie_expropriere', ''))
        p_dec = normalize(pdf.get('decizie_expropriere_pdf', ''))
        # Extract just numbers for comparison
        m_nums = re.findall(r'\d+', m_dec)
        p_nums = re.findall(r'\d+', p_dec)
        if m_nums and p_nums and m_nums != p_nums:
            issues.append({
                'camp': 'Decizie expropriere',
                'master': master.get('decizie_expropriere', ''),
                'pdf': pdf.get('decizie_expropriere_pdf', ''),
            })

    # HG
    if pdf.get('hg_pdf'):
        m_hg = re.sub(r'\s', '', str(master.get('hg', '')))
        p_hg = re.sub(r'\s', '', str(pdf.get('hg_pdf', '')))
        if m_hg and p_hg and m_hg != p_hg:
            issues.append({
                'camp': 'HG',
                'master': master.get('hg', ''),
                'pdf': pdf.get('hg_pdf', ''),
            })

    return issues


def process_all(master_path, pdf_folder, progress_callback=None):
    """Proceseaza MASTER + toate PDF-urile. Returneaza rezultate."""
    master_entries = read_master(master_path)
    master_by_pv = {e['nr_pv']: e for e in master_entries}

    pdf_files = sorted([f for f in os.listdir(pdf_folder) if f.lower().endswith('.pdf')],
                       key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else 0)

    results = []
    for idx, pdf_file in enumerate(pdf_files):
        if progress_callback:
            progress_callback(idx + 1, len(pdf_files), pdf_file)

        pdf_path = os.path.join(pdf_folder, pdf_file)
        text = ocr_pdf(pdf_path)
        pdf_data = extract_data_from_pdf(text, pdf_file)

        nr_pv = pdf_data.get('nr_pv')
        if nr_pv and nr_pv in master_by_pv:
            master_entry = master_by_pv[nr_pv]
            issues = compare_entry(master_entry, pdf_data)
            results.append({
                'pdf_file': pdf_file,
                'nr_pv': nr_pv,
                'uat': master_entry.get('uat', ''),
                'pozitie_hg': master_entry.get('pozitie_hg', ''),
                'proprietar': master_entry.get('nume1', ''),
                'issues': issues,
                'pdf_data': pdf_data,
                'status': 'NECONCORDANȚE' if issues else 'OK',
            })
        else:
            results.append({
                'pdf_file': pdf_file,
                'nr_pv': nr_pv,
                'uat': pdf_data.get('uat', ''),
                'pozitie_hg': pdf_data.get('pozitie_hg', ''),
                'proprietar': pdf_data.get('proprietar_pdf', ''),
                'issues': [{'camp': 'Nr. PV', 'master': 'NEGĂSIT', 'pdf': str(nr_pv)}],
                'pdf_data': pdf_data,
                'status': 'NEGĂSIT ÎN MASTER',
            })

    return results, master_entries


def generate_report(results, output_path):
    """Genereaza raport Excel cu neconcordantele."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Neconcordanțe"

    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='C00000')
    df = Font(name='Arial', size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('A1:G1')
    ws['A1'] = 'RAPORT NECONCORDANȚE - VERIFICARE H SI PV vs MASTER'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='C00000')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Nr.\ncrt.', 'Fișier PDF', 'Nr.\nPV/HSD', 'UAT', 'Poziție\nHG', 'Câmp cu\nneconcordanță', 'Valoare\nMASTER', 'Valoare\nPDF']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ca
        cell.border = tb

    row = 4
    crt = 0
    issues_only = [r for r in results if r['issues']]
    for r in issues_only:
        for issue in r['issues']:
            crt += 1
            vals = [crt, r['pdf_file'], r['nr_pv'], r.get('uat', ''),
                    r.get('pozitie_hg', ''), issue['camp'], issue['master'], issue['pdf']]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = df
                cell.alignment = ca if col <= 5 else Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = tb
            row += 1

    # Summary sheet
    ws2 = wb.create_sheet("Sumar")
    ws2['A1'] = 'SUMAR VERIFICARE'
    ws2['A1'].font = Font(name='Arial', bold=True, size=14)
    ws2['A3'] = 'Total PDF-uri verificate:'
    ws2['B3'] = len(results)
    ws2['A4'] = 'PDF-uri OK (fără neconcordanțe):'
    ws2['B4'] = sum(1 for r in results if not r['issues'])
    ws2['A5'] = 'PDF-uri cu neconcordanțe:'
    ws2['B5'] = sum(1 for r in results if r['issues'])
    ws2['A5'].font = Font(name='Arial', bold=True, color='FF0000')
    ws2['B5'].font = Font(name='Arial', bold=True, color='FF0000')
    ws2['A6'] = 'Total neconcordanțe găsite:'
    ws2['B6'] = sum(len(r['issues']) for r in results)

    for c in ['A', 'B']:
        ws2.column_dimensions[c].width = 35

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 30
    ws.row_dimensions[3].height = 40

    wb.save(output_path)
    return output_path
